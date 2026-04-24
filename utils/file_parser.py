import pandas as pd
import io

# ─── Template CSV per Credit Scoring ───────────────────────────────────────────
CREDIT_TEMPLATE = """campo,valore,descrizione
nome_azienda,,Nome dell'azienda
ebit,,EBIT - Reddito Operativo (EUR)
ammortamenti,,Ammortamenti e svalutazioni (EUR)
interessi_passivi,,Interessi passivi (EUR)
rimborso_debito,,Rimborso debito annuo (EUR)
debiti_finanziari_totali,,Debiti finanziari totali (EUR)
patrimonio_netto,,Patrimonio Netto (EUR)
ebitda,,EBITDA (EUR)
ricavi_netti,,Ricavi Netti (EUR)
attivo_corrente,,Attivo Corrente (EUR)
passivo_corrente,,Passivo Corrente (EUR)
"""

# ─── Template CSV per Z-Score ───────────────────────────────────────────────────
ZSCORE_TEMPLATE = """campo,valore,descrizione
nome_azienda,,Nome dell'azienda
totale_attivo,,Totale Attivo (EUR)
utili_non_distribuiti,,Utili non distribuiti / Riserve (EUR)
ebit,,EBIT - Reddito Operativo (EUR)
capitale_circolante_netto,,Capitale Circolante Netto = Attivo Corr - Passivo Corr (EUR)
totale_passivita,,Totale Passivita (EUR)
patrimonio_netto,,Patrimonio Netto Contabile (EUR)
ricavi_netti,,Ricavi Netti (EUR)
"""

# ─── Template CSV per Financial Ratios ──────────────────────────────────────────
RATIOS_TEMPLATE = """campo,valore,descrizione
nome_azienda,,Nome dell'azienda
total_assets,,Totale Attivo (EUR)
total_liabilities,,Totale Passivo (EUR)
equity,,Patrimonio Netto (EUR)
current_assets,,Attivo Corrente (EUR)
current_liabilities,,Passivo Corrente (EUR)
revenue,,Ricavi Netti (EUR)
ebit,,EBIT - Reddito Operativo (EUR)
ebitda,,EBITDA (EUR)
net_income,,Utile Netto (EUR)
depreciation,,Ammortamenti (EUR)
interest_expense,,Oneri Finanziari (EUR)
inventory,,Magazzino (EUR)
accounts_receivable,,Crediti Commerciali (EUR)
accounts_payable,,Debiti Commerciali (EUR)
retained_earnings,,Utili portati a nuovo (EUR)
cash,,Cassa e disponibilita liquide (EUR)
long_term_debt,,Debiti a lungo termine (EUR)
"""


def _safe_float(val, default=0.0):
    try:
        return float(str(val).replace(",", ".").replace("EUR", "").replace("€", "").replace(" ", "").strip()) if val else default
    except:
        return default


def _parse_dataframe(df):
    """Converte un DataFrame (campo/valore) o riga-singola in dict."""
    if "campo" in df.columns and "valore" in df.columns:
        return dict(zip(df["campo"].str.strip().str.lower(), df["valore"]))
    elif len(df) >= 1:
        return {col.strip().lower(): df[col].iloc[0] for col in df.columns}
    return {}


def _read_file(uploaded_file):
    """Legge CSV o Excel e restituisce DataFrame."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)


def parse_credit_file(uploaded_file):
    """Parsa file CSV/Excel per Credit Scoring. Restituisce dict o errore."""
    try:
        df = _read_file(uploaded_file)
        data = _parse_dataframe(df)
        return {
            "success": True,
            "nome_azienda": str(data.get("nome_azienda", "")),
            "ebit": _safe_float(data.get("ebit")),
            "depreciation": _safe_float(data.get("ammortamenti")),
            "interest_expense": _safe_float(data.get("interessi_passivi")),
            "debt_repayment": _safe_float(data.get("rimborso_debito")),
            "total_debt": _safe_float(data.get("debiti_finanziari_totali")),
            "total_equity": _safe_float(data.get("patrimonio_netto"), 1.0),
            "ebitda": _safe_float(data.get("ebitda")),
            "net_revenue": _safe_float(data.get("ricavi_netti"), 1.0),
            "current_assets": _safe_float(data.get("attivo_corrente")),
            "current_liabilities": _safe_float(data.get("passivo_corrente"), 1.0),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def parse_zscore_file(uploaded_file):
    """Parsa file CSV/Excel per Z-Score. Restituisce dict o errore."""
    try:
        df = _read_file(uploaded_file)
        data = _parse_dataframe(df)
        return {
            "success": True,
            "nome_azienda": str(data.get("nome_azienda", "")),
            "total_assets": _safe_float(data.get("totale_attivo"), 1.0),
            "retained_earnings": _safe_float(data.get("utili_non_distribuiti")),
            "ebit": _safe_float(data.get("ebit")),
            "working_capital": _safe_float(data.get("capitale_circolante_netto")),
            "total_liabilities": _safe_float(data.get("totale_passivita"), 1.0),
            "equity_input": _safe_float(data.get("patrimonio_netto")),
            "revenue": _safe_float(data.get("ricavi_netti")),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def parse_ratios_file(uploaded_file):
    """Parsa file CSV/Excel per Financial Ratios. Restituisce dict o errore."""
    try:
        df = _read_file(uploaded_file)
        data = _parse_dataframe(df)
        keys = [
            "total_assets", "total_liabilities", "equity", "current_assets",
            "current_liabilities", "revenue", "ebit", "ebitda", "net_income",
            "depreciation", "interest_expense", "inventory", "accounts_receivable",
            "accounts_payable", "retained_earnings", "cash", "long_term_debt",
        ]
        result = {
            "success": True,
            "nome_azienda": str(data.get("nome_azienda", "")),
        }
        for k in keys:
            result[k] = _safe_float(data.get(k))
        return result
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_credit_template_bytes():
    return CREDIT_TEMPLATE.encode("utf-8")


def get_zscore_template_bytes():
    return ZSCORE_TEMPLATE.encode("utf-8")


def get_ratios_template_bytes():
    return RATIOS_TEMPLATE.encode("utf-8")


def get_credit_template_excel():
    """Genera template Excel per Credit Scoring."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Credit Scoring"
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 45
        rows = [
            ("campo", "valore", "descrizione"),
            ("nome_azienda", "", "Nome dell'azienda"),
            ("ebit", "", "EBIT - Reddito Operativo (EUR)"),
            ("ammortamenti", "", "Ammortamenti e svalutazioni (EUR)"),
            ("interessi_passivi", "", "Interessi passivi (EUR)"),
            ("rimborso_debito", "", "Rimborso debito annuo (EUR)"),
            ("debiti_finanziari_totali", "", "Debiti finanziari totali (EUR)"),
            ("patrimonio_netto", "", "Patrimonio Netto (EUR)"),
            ("ebitda", "", "EBITDA (EUR)"),
            ("ricavi_netti", "", "Ricavi Netti (EUR)"),
            ("attivo_corrente", "", "Attivo Corrente (EUR)"),
            ("passivo_corrente", "", "Passivo Corrente (EUR)"),
        ]
        for i, row in enumerate(rows, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                elif j == 2:
                    cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except:
        return get_credit_template_bytes()


def get_zscore_template_excel():
    """Genera template Excel per Z-Score."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Z-Score"
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 55
        rows = [
            ("campo", "valore", "descrizione"),
            ("nome_azienda", "", "Nome dell'azienda"),
            ("totale_attivo", "", "Totale Attivo (EUR)"),
            ("utili_non_distribuiti", "", "Utili non distribuiti / Riserve (EUR)"),
            ("ebit", "", "EBIT - Reddito Operativo (EUR)"),
            ("capitale_circolante_netto", "", "Capitale Circolante Netto = Attivo Corr - Passivo Corr (EUR)"),
            ("totale_passivita", "", "Totale Passivita (EUR)"),
            ("patrimonio_netto", "", "Patrimonio Netto Contabile (EUR)"),
            ("ricavi_netti", "", "Ricavi Netti (EUR)"),
        ]
        for i, row in enumerate(rows, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                elif j == 2:
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except:
        return get_zscore_template_bytes()


def get_ratios_template_excel():
    """Genera template Excel per Financial Ratios."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Ratios"
        header_fill = PatternFill(start_color="006064", end_color="006064", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 45
        rows = [
            ("campo", "valore", "descrizione"),
            ("nome_azienda", "", "Nome dell'azienda"),
            ("total_assets", "", "Totale Attivo (EUR)"),
            ("total_liabilities", "", "Totale Passivo (EUR)"),
            ("equity", "", "Patrimonio Netto (EUR)"),
            ("current_assets", "", "Attivo Corrente (EUR)"),
            ("current_liabilities", "", "Passivo Corrente (EUR)"),
            ("revenue", "", "Ricavi Netti (EUR)"),
            ("ebit", "", "EBIT - Reddito Operativo (EUR)"),
            ("ebitda", "", "EBITDA (EUR)"),
            ("net_income", "", "Utile Netto (EUR)"),
            ("depreciation", "", "Ammortamenti (EUR)"),
            ("interest_expense", "", "Oneri Finanziari (EUR)"),
            ("inventory", "", "Magazzino (EUR)"),
            ("accounts_receivable", "", "Crediti Commerciali (EUR)"),
            ("accounts_payable", "", "Debiti Commerciali (EUR)"),
            ("retained_earnings", "", "Utili portati a nuovo (EUR)"),
            ("cash", "", "Cassa e disponibilita liquide (EUR)"),
            ("long_term_debt", "", "Debiti a lungo termine (EUR)"),
        ]
        for i, row in enumerate(rows, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                elif j == 2:
                    cell.fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except:
        return get_ratios_template_bytes()


def generate_template_csv():
    """Alias per compatibilita."""
    return get_zscore_template_bytes()
